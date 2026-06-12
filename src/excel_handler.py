"""
Excel handler for BlogMaker.

Reads pending topics from and writes status updates to the topics spreadsheet.
Uses openpyxl for .xlsx manipulation.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from src.logger import get_logger
from src.models import TopicRow

logger = get_logger("excel")

# Expected column positions (1-indexed)
COL_TOPIC = 1
COL_STATUS = 2
COL_PRIORITY = 3
HEADER_ROW = 1


def read_pending_topic(filepath: str) -> TopicRow | None:
    """
    Read the first pending topic from the Excel spreadsheet.

    Scans rows from top to bottom and returns the first row
    where Status == "Pending" (case-insensitive).

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        TopicRow if a pending topic is found, None otherwise.

    Raises:
        FileNotFoundError: If the Excel file does not exist.
        ValueError: If the spreadsheet is empty or has no data rows.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(
            f"Topics file not found: {filepath}\n"
            "Create a topics.xlsx file with columns: Topic | Status | Priority"
        )

    wb: Workbook = load_workbook(filepath, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError(f"No active worksheet found in {filepath}")

    # Check for data rows (beyond header)
    if ws.max_row is None or ws.max_row < 2:
        logger.info("No data rows found in %s", filepath)
        wb.close()
        return None

    logger.info("Scanning %s for pending topics (%d rows)...", filepath, ws.max_row - 1)

    for row_num in range(HEADER_ROW + 1, ws.max_row + 1):
        topic_cell = ws.cell(row=row_num, column=COL_TOPIC).value
        status_cell = ws.cell(row=row_num, column=COL_STATUS).value
        priority_cell = ws.cell(row=row_num, column=COL_PRIORITY).value

        # Skip empty rows
        if not topic_cell:
            continue

        status = str(status_cell).strip() if status_cell else ""

        if status.lower() == "pending":
            topic = str(topic_cell).strip()
            priority = str(priority_cell).strip() if priority_cell else "Medium"

            logger.info(
                "Found pending topic: '%s' (Priority: %s, Row: %d)",
                topic, priority, row_num,
            )
            wb.close()
            return TopicRow(
                topic=topic,
                status=status,
                priority=priority,
                row_number=row_num,
            )

    logger.info("No pending topics found in %s", filepath)
    wb.close()
    return None


def mark_topic_done(filepath: str, row_number: int) -> None:
    """
    Mark a topic row as Done in the Excel spreadsheet.

    Updates only the Status cell for the specified row.
    Preserves all other data and formatting.

    Args:
        filepath: Path to the .xlsx file.
        row_number: The 1-indexed row number to update.

    Raises:
        FileNotFoundError: If the Excel file does not exist.
        ValueError: If the row number is invalid.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Topics file not found: {filepath}")

    wb: Workbook = load_workbook(filepath)
    ws = wb.active

    if ws is None:
        raise ValueError(f"No active worksheet found in {filepath}")

    if row_number < 2 or (ws.max_row is not None and row_number > ws.max_row):
        raise ValueError(
            f"Invalid row number: {row_number}. "
            f"Valid range: 2–{ws.max_row}"
        )

    # Read current topic for logging
    current_topic = ws.cell(row=row_number, column=COL_TOPIC).value

    # Update status
    ws.cell(row=row_number, column=COL_STATUS, value="Done")
    wb.save(filepath)
    wb.close()

    logger.info(
        "Marked topic '%s' (row %d) as Done in %s",
        current_topic, row_number, filepath,
    )


def create_example_xlsx(filepath: str) -> None:
    """
    Create an example topics.xlsx file with sample data.

    Args:
        filepath: Path to create the .xlsx file at.
    """
    from openpyxl import Workbook as NewWorkbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = NewWorkbook()
    ws = wb.active
    ws.title = "Topics"

    # Header styling
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Write headers
    headers = ["Topic", "Status", "Priority"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Sample data
    topics = [
        ("Banking with AI", "Pending", "High"),
        ("Future of Jobs", "Pending", "Medium"),
        ("Quantum Computing in Finance", "Pending", "High"),
        ("Green Energy Policy 2025", "Pending", "Low"),
        ("Cybersecurity Trends", "Pending", "Medium"),
    ]

    for row_idx, (topic, status, priority) in enumerate(topics, start=2):
        ws.cell(row=row_idx, column=COL_TOPIC, value=topic)
        ws.cell(row=row_idx, column=COL_STATUS, value=status)
        ws.cell(row=row_idx, column=COL_PRIORITY, value=priority)

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12

    wb.save(filepath)
    wb.close()

    logger.info("Created example topics file: %s", filepath)
